# -*- coding: utf-8 -*-
"""
Canopy Pricing Configurator — Streamlit app (v13 — quote-type & double-canopy modes)

v13 spec (2026-05-15) — see canopy_pricing_context.md Session 2026-05-15 entry.
Built on v12 (flood lights, light-count override, flat light pricing, large-canopy GP).

v13 additions:
 - Imaging disclaimer in CANOPY BRANDING / IMAGING sections (when WE supply).
 - Quote Type selector at top: Imaging Only | Canopies.
 - Canopies → Single Canopy | Double Canopy (Gas + Diesel).
 - Imaging-only flow: brand + dispensers + labor days + misc labor $ + MID 4-way
   with USER-INPUT MID labor amount (no MID_Labor lookup).
 - Per-canopy independence in double mode: branding, GP, dispensers, dimensions,
   lights are all per-canopy. Shared: customer info, MID, removal, distance,
   tax rate.
 - Canopy 2 GP defaults to canopy 1 GP − 3% (password-overridable).
 - Filename: `Q-YYYYMMDD-NNN (StreetName).docx`. Imaging-only adds " - Imaging Only".
 - load_config now reads ALL Brand_Imaging rows (until "Note:" row).
"""

import csv
import datetime as dt
import io
import math
import os
import tempfile
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from proposal_writer import build_proposal

# Google Sheets backend (production). Falls back to local CSV when creds aren't
# configured (e.g., local dev), so the app still runs without google deps.
try:
    import gspread  # pyrefly: ignore[missing-import]
    from google.oauth2.service_account import Credentials  # pyrefly: ignore[missing-import]

    _HAS_GSPREAD = True
except ImportError:
    _HAS_GSPREAD = False

CONFIG_PATH = Path(__file__).parent / "canopy_config.xlsx"
APEC_LOGO = Path(__file__).parent / "Apec Imaging Logo.jpg"
GEO_LOGO = Path(__file__).parent / "GEO Canopies logo.jpg"
PASSWORD = "cheap"
ADMIN_PASSWORD = "profit_tracker"

# v13: GP reduction applied to the second canopy in double mode.
DOUBLE_CANOPY_GP_REDUCTION = 0.03

# ─── Quote tracker storage ─────────────────────────────────────────────────
SHEET_ID = "1pnfCv70Y4UBWw8A2Yi49QBvr2ErwRDy7raxPdkGkC1w"
WORKSHEET_NAME = "Tracker"
TRACKER_PATH = Path(__file__).parent / "quote_tracker.csv"
TRACKER_COLUMNS = [
    "Date",
    "Quote No",
    "Customer Name",
    "City",
    "Sales Rep",
    "Grand Total",
    "Profitability",
]
GSHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


@st.cache_resource(show_spinner=False)
def _get_tracker_sheet():
    if not _HAS_GSPREAD:
        return None
    try:
        creds_dict = dict(st.secrets["gcp_service_account"])
    except Exception:
        return None
    try:
        creds = Credentials.from_service_account_info(creds_dict, scopes=GSHEETS_SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SHEET_ID)
        try:
            ws = sh.worksheet(WORKSHEET_NAME)
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(
                title=WORKSHEET_NAME, rows=1000, cols=len(TRACKER_COLUMNS)
            )
            ws.append_row(TRACKER_COLUMNS, value_input_option="USER_ENTERED")
        if not ws.row_values(1):
            ws.append_row(TRACKER_COLUMNS, value_input_option="USER_ENTERED")
        return ws
    except Exception as e:
        st.warning(
            f"Google Sheets tracker unavailable ({e}); using local CSV fallback."
        )
        return None


US_STATES = [
    "AL",
    "AK",
    "AZ",
    "AR",
    "CA",
    "CO",
    "CT",
    "DE",
    "FL",
    "GA",
    "HI",
    "ID",
    "IL",
    "IN",
    "IA",
    "KS",
    "KY",
    "LA",
    "ME",
    "MD",
    "MA",
    "MI",
    "MN",
    "MS",
    "MO",
    "MT",
    "NE",
    "NV",
    "NH",
    "NJ",
    "NM",
    "NY",
    "NC",
    "ND",
    "OH",
    "OK",
    "OR",
    "PA",
    "RI",
    "SC",
    "SD",
    "TN",
    "TX",
    "UT",
    "VT",
    "VA",
    "WA",
    "WV",
    "WI",
    "WY",
    "DC",
]


@st.cache_data(ttl=60)
def load_config():
    wb = load_workbook(CONFIG_PATH, data_only=True)
    c = {}

    def _num(cell, fallback):
        v = cell.value if hasattr(cell, "value") else cell
        return v if isinstance(v, (int, float)) and v is not False else fallback

    s = wb["Settings"]
    settings_map = {}
    for r in range(3, s.max_row + 1):
        key = s.cell(r, 1).value
        if key:
            settings_map[str(key)] = (s.cell(r, 2).value, s.cell(r, 3).value)

    def _sval(name, fallback=None):
        v, _ = settings_map.get(name, (fallback, None))
        return v if v is not None else fallback

    def _scost(name, fallback=None):
        _, v = settings_map.get(name, (None, fallback))
        return v if v is not None else fallback

    c["canopy_height_default_ft"] = _sval("canopy_height_default_ft", 16.5)
    c["default_dispensers"] = _sval("default_dispensers", 4)
    c["acm_panel_width_ft"] = _sval("acm_panel_width_ft", 10)
    c["tax_default_rate"] = _sval("tax_default_rate", 0.0825)
    c["laborer_count"] = _sval("laborer_count", 4)
    c["hours_per_day"] = _sval("hours_per_day", 10)
    c["hourly_rate"] = _sval("hourly_rate", 40)
    raw = _sval("labor_daily_rate")
    c["labor_daily_rate"] = (
        raw
        if raw is not None
        else (c["laborer_count"] * c["hours_per_day"] * c["hourly_rate"])
    )
    c["branded_labor_adder"] = _sval("branded_labor_adder", 3500)
    c["two_disp_labor_adder"] = _sval("two_disp_labor_adder", 3500)
    c["shipping_handling"] = _sval("shipping_handling", 3500)
    c["branded_labor_adder_cost"] = _scost(
        "branded_labor_adder", c["branded_labor_adder"]
    )
    c["two_disp_labor_adder_cost"] = _scost(
        "two_disp_labor_adder", c["two_disp_labor_adder"]
    )
    c["shipping_handling_cost"] = _scost("shipping_handling", c["shipping_handling"])
    raw_lcost = _scost("labor_daily_rate")
    c["labor_daily_rate_cost"] = (
        raw_lcost if raw_lcost is not None else c["labor_daily_rate"]
    )

    c["gas_default_left_overhang_ft"] = _sval("gas_default_left_overhang_ft", 10)
    c["gas_default_right_overhang_ft"] = _sval("gas_default_right_overhang_ft", 10)
    c["gas_default_dispenser_spacing_ft"] = _sval(
        "gas_default_dispenser_spacing_ft", 28
    )
    c["gas_default_depth_ft"] = _sval("gas_default_depth_ft", 24)
    c["diesel_default_left_overhang_ft"] = _sval("diesel_default_left_overhang_ft", 2)
    c["diesel_default_right_overhang_ft"] = _sval("diesel_default_right_overhang_ft", 2)
    c["diesel_default_dispenser_spacing_ft"] = _sval(
        "diesel_default_dispenser_spacing_ft", 24
    )
    c["diesel_default_depth_ft"] = _sval("diesel_default_depth_ft", 20)
    c["distance_threshold_miles"] = _sval("distance_threshold_miles", 60)

    m = wb["Material_Rates"]
    mat_map = {}
    for r in range(3, m.max_row + 1):
        key = m.cell(r, 1).value
        if key:
            mat_map[str(key)] = (m.cell(r, 2).value, m.cell(r, 3).value)

    def _mval(name, fallback=None):
        v, _ = mat_map.get(name, (fallback, None))
        return v if v is not None else fallback

    def _mcost(name, fallback=None):
        _, v = mat_map.get(name, (None, fallback))
        return v if v is not None else fallback

    c["acm_per_panel"] = _mval("acm_per_panel", 550)
    c["steel_primary"] = _mval("steel_primary_per_col", 4000)
    c["steel_secondary"] = _mval("steel_secondary_per_col", 2000)
    c["decking_per_col"] = _mval("decking_per_col", 5000)
    c["light_unit_price_standard"] = _mval("light_unit_price_standard", 150)
    c["light_unit_price_lsi"] = _mval("light_unit_price_lsi", 275)
    c["acm_per_panel_cost"] = _mcost("acm_per_panel", c["acm_per_panel"])
    c["steel_primary_cost"] = _mcost("steel_primary_per_col", c["steel_primary"])
    c["steel_secondary_cost"] = _mcost("steel_secondary_per_col", c["steel_secondary"])
    c["decking_per_col_cost"] = _mcost("decking_per_col", c["decking_per_col"])
    c["light_unit_price_standard_cost"] = _mcost(
        "light_unit_price_standard", c["light_unit_price_standard"]
    )
    c["light_unit_price_lsi_cost"] = _mcost(
        "light_unit_price_lsi", c["light_unit_price_lsi"]
    )
    c["flood_light_unit_price"] = _mval("flood_light_unit_price", 450)
    c["flood_light_unit_price_cost"] = _mcost("flood_light_unit_price", 350)
    c["distance_adder_per_install_day"] = _mval("distance_adder_per_install_day", 1000)
    c["distance_adder_per_install_day_cost"] = _mcost(
        "distance_adder_per_install_day", 600
    )
    c["demo_default_retail"] = _mval("demo_default", 8000)
    c["demo_default_cost"] = _mcost("demo_default", 5000)

    mt = wb["MISC_Tiers"]
    c["misc_tiers"] = [
        (mt.cell(r, 1).value, mt.cell(r, 2).value, mt.cell(r, 3).value)
        for r in range(3, 6)
        if mt.cell(r, 1).value is not None
    ]
    c["misc_tiers_cost"] = [
        (
            mt.cell(r, 1).value,
            mt.cell(r, 2).value,
            _num(mt.cell(r, 4), mt.cell(r, 3).value),
        )
        for r in range(3, 6)
        if mt.cell(r, 1).value is not None
    ]

    ld = wb["Labor_Days"]
    c["labor_days"] = {}
    for r in range(3, 12):
        d, days = ld.cell(r, 1).value, ld.cell(r, 2).value
        if d is not None and days is not None:
            c["labor_days"][int(d)] = int(days)

    # v13: read ALL brand rows (was hard-coded to rows 3-9). Stop at "Note:" row
    # or first empty brand cell after the header gap.
    bi = wb["Brand_Imaging"]
    c["brand_imaging"] = {}
    for r in range(3, bi.max_row + 2):
        brand = bi.cell(r, 1).value
        if brand is None:
            continue
        bs = str(brand).strip()
        if bs.startswith("Note") or not bs:
            break
        prices = {}
        for disp in range(2, 11):
            v = bi.cell(r, disp).value
            if v is not None:
                prices[disp] = v
        c["brand_imaging"][bs] = prices

    mp = wb["MID_PriceSign"]
    c["mid_prices"] = {}
    c["mid_prices_cost"] = {}
    for r in range(3, 8):
        brand, price = mp.cell(r, 1).value, mp.cell(r, 2).value
        if brand and price is not None:
            c["mid_prices"][brand] = price
            c["mid_prices_cost"][brand] = _num(mp.cell(r, 3), price)

    ml = wb["MID_Labor"]
    c["mid_labor_by_company"] = {}
    c["mid_labor_by_company_cost"] = {}
    for r in range(3, ml.max_row + 1):
        code, amt = ml.cell(r, 1).value, ml.cell(r, 2).value
        if code and amt is not None:
            c["mid_labor_by_company"][str(code)] = amt
            c["mid_labor_by_company_cost"][str(code)] = _num(ml.cell(r, 3), amt)

    sp = wb["Salespeople"]
    c["salespeople"] = {}
    for r in range(3, 6):
        name = sp.cell(r, 1).value
        if name:
            c["salespeople"][name] = {
                "phone": sp.cell(r, 2).value,
                "email": sp.cell(r, 3).value,
            }

    ipl = wb["Internal_PL"]
    c["gp_apec"] = ipl["B6"].value
    c["gp_geo"] = ipl["B7"].value
    return c


def auto_dimensions(
    canopy_type, dispensers, left_overhang, right_overhang, spacing, fuel_type, cfg
):
    n_per_row = dispensers // 2 if canopy_type == "Stacked" else dispensers
    width = left_overhang + (n_per_row - 1) * spacing + right_overhang
    base_depth_key = (
        "gas_default_depth_ft" if fuel_type == "Gas" else "diesel_default_depth_ft"
    )
    base_depth = cfg.get(base_depth_key, 24)
    depth = base_depth * 2 if canopy_type == "Stacked" else base_depth
    return width, depth


def column_count(canopy_type, dispensers, double_col):
    return 2 * dispensers if (canopy_type == "Dive-in" and double_col) else dispensers


def default_light_count(canopy_type, dispensers, double_col):
    cols = column_count(canopy_type, dispensers, double_col)
    if canopy_type == "Dive-in" and double_col:
        return cols
    return cols * 4


def compute_acm(width, depth, cfg):
    a1 = 2 * width + 2 * depth
    a2 = math.ceil(a1 / cfg["acm_panel_width_ft"])
    return a2 * cfg["acm_per_panel"]


def lookup_misc(columns, cfg):
    for lo, hi, price in cfg["misc_tiers"]:
        if lo <= columns <= hi:
            return price
    return 0


def lookup_misc_cost(columns, cfg):
    for lo, hi, cost in cfg["misc_tiers_cost"]:
        if lo <= columns <= hi:
            return cost
    return 0


def compute_quote(
    *,
    dispensers,
    canopy_type,
    double_col,
    fuel_type,
    branded,
    brand_name,
    customer_supplied_imaging,
    include_mid_material,
    mid_brand,
    include_mid_labor,
    mid_labor_amt,
    left_overhang,
    right_overhang,
    spacing,
    width,
    depth,
    distance_miles,
    include_demo,
    demo_retail,
    demo_cost,
    tax_rate,
    gp_rate,
    light_type,
    light_count,
    flood_light_count,
    flood_light_unit_retail,
    company_key,
    cfg,
):
    """v13: signature unchanged from v12. compute one canopy."""
    columns = column_count(canopy_type, dispensers, double_col)
    items_not_included = []

    acm = 0 if branded else compute_acm(width, depth, cfg)
    steel_primary = dispensers * cfg["steel_primary"]
    decking = columns * cfg["decking_per_col"]
    if light_type == "Customer-provided":
        light_unit_retail, light_unit_cost = 0, 0
        eff_light_count = 0
    elif light_type == "LSI":
        light_unit_retail = cfg["light_unit_price_lsi"]
        light_unit_cost = cfg["light_unit_price_lsi_cost"]
        eff_light_count = light_count
    else:
        light_unit_retail = cfg["light_unit_price_standard"]
        light_unit_cost = cfg["light_unit_price_standard_cost"]
        eff_light_count = light_count
    lights = eff_light_count * light_unit_retail
    lights_cost = eff_light_count * light_unit_cost
    flood_lights = flood_light_count * flood_light_unit_retail
    flood_lights_cost = flood_light_count * cfg["flood_light_unit_price_cost"]
    misc = lookup_misc(columns, cfg)
    marked_up_material = acm + steel_primary + decking + misc

    steel_secondary = (
        dispensers * cfg["steel_secondary"]
        if (canopy_type == "Dive-in" and double_col)
        else 0
    )

    brand_imaging_amt = 0
    if branded and not customer_supplied_imaging:
        prices = cfg["brand_imaging"].get(brand_name, {})
        if dispensers in prices:
            brand_imaging_amt = prices[dispensers]
        else:
            items_not_included.append(
                f"Brand imaging price for {brand_name} at {dispensers} dispensers "
                f"is not on file in this configurator — verify before quoting."
            )

    if not branded:
        items_not_included.append("Brand imaging is not included.")

    shipping = (
        cfg["shipping_handling"] if (branded and not customer_supplied_imaging) else 0
    )

    mid_material = 0
    if include_mid_material:
        if mid_brand and mid_brand in cfg["mid_prices"]:
            mid_material = cfg["mid_prices"][mid_brand]
        else:
            items_not_included.append(
                f"Price sign material for {mid_brand} not in cheat sheet."
            )

    at_cost_material = steel_secondary + shipping + brand_imaging_amt + mid_material

    days = cfg["labor_days"].get(int(dispensers), 0)
    base_labor = days * cfg["labor_daily_rate"]
    marked_up_labor = base_labor

    branded_add = cfg["branded_labor_adder"] if branded else 0
    two_disp_add = cfg["two_disp_labor_adder"] if dispensers == 2 else 0
    mid_labor_total = mid_labor_amt if include_mid_labor else 0

    threshold = cfg.get("distance_threshold_miles", 60)
    if distance_miles is not None and distance_miles > threshold:
        travel_adder_retail = days * cfg.get("distance_adder_per_install_day", 0)
        travel_adder_cost = days * cfg.get("distance_adder_per_install_day_cost", 0)
    else:
        travel_adder_retail = 0
        travel_adder_cost = 0

    at_cost_labor = branded_add + two_disp_add + mid_labor_total + travel_adder_retail

    retail_material = (
        marked_up_material * (1 + gp_rate) + at_cost_material + lights + flood_lights
    )
    retail_labor = marked_up_labor * (1 + gp_rate) + at_cost_labor
    demo_retail_eff = demo_retail if include_demo else 0
    demo_cost_eff = demo_cost if include_demo else 0
    tax = retail_material * tax_rate

    material_cost = marked_up_material + at_cost_material + lights + flood_lights
    labor_cost = marked_up_labor + at_cost_labor
    final = retail_material + retail_labor + demo_retail_eff + tax

    acm_cost = (
        0
        if branded
        else compute_acm(width, depth, cfg)
        * (
            cfg["acm_per_panel_cost"] / cfg["acm_per_panel"]
            if cfg["acm_per_panel"]
            else 0
        )
    )
    steel_primary_cost = dispensers * cfg["steel_primary_cost"]
    decking_cost = columns * cfg["decking_per_col_cost"]
    misc_cost = lookup_misc_cost(columns, cfg)
    steel_secondary_cost = (
        dispensers * cfg["steel_secondary_cost"]
        if (canopy_type == "Dive-in" and double_col)
        else 0
    )
    shipping_cost = (
        cfg["shipping_handling_cost"]
        if (branded and not customer_supplied_imaging)
        else 0
    )
    brand_imaging_cost = brand_imaging_amt
    mid_material_cost = (
        cfg["mid_prices_cost"].get(mid_brand, mid_material)
        if include_mid_material and mid_brand
        else 0
    )
    base_labor_cost = days * cfg["labor_daily_rate_cost"]
    branded_labor_add_cost = cfg["branded_labor_adder_cost"] if branded else 0
    two_disp_labor_add_cost = cfg["two_disp_labor_adder_cost"] if dispensers == 2 else 0
    mid_labor_cost = mid_labor_amt if include_mid_labor else 0

    true_cost_material = (
        acm_cost
        + steel_primary_cost
        + decking_cost
        + lights_cost
        + flood_lights_cost
        + misc_cost
        + steel_secondary_cost
        + shipping_cost
        + brand_imaging_cost
        + mid_material_cost
    )
    true_cost_labor = (
        base_labor_cost
        + branded_labor_add_cost
        + two_disp_labor_add_cost
        + mid_labor_cost
        + travel_adder_cost
        + demo_cost_eff
    )
    true_cost_total = true_cost_material + true_cost_labor

    revenue = retail_material + retail_labor + demo_retail_eff
    profit = revenue - true_cost_total

    return {
        "material_cost": material_cost,
        "labor_cost": labor_cost,
        "cost_total": material_cost + labor_cost + demo_cost_eff,
        "retail_material": retail_material,
        "retail_labor": retail_labor,
        "retail_total": retail_material + retail_labor + demo_retail_eff,
        "tax": tax,
        "final": final,
        "columns": columns,
        "labor_days": days,
        "items_not_included": items_not_included,
        "acm": acm,
        "steel": steel_primary,
        "decking": decking,
        "lights": lights,
        "flood_lights": flood_lights,
        "light_type": light_type,
        "light_count": light_count,
        "flood_light_count": flood_light_count,
        "flood_light_unit_retail": flood_light_unit_retail,
        "misc": misc,
        "steel_secondary": steel_secondary,
        "shipping": shipping,
        "brand_imaging": brand_imaging_amt,
        "mid_material": mid_material,
        "mid_labor": mid_labor_total,
        "base_labor": base_labor,
        "branded_labor_add": branded_add,
        "two_disp_labor_add": two_disp_add,
        "travel_adder_retail": travel_adder_retail,
        "travel_adder_cost": travel_adder_cost,
        "demo_retail": demo_retail_eff,
        "demo_cost": demo_cost_eff,
        "revenue": revenue,
        "true_cost": true_cost_total,
        "profit": profit,
    }


def compute_imaging_only(
    *,
    brand_name,
    dispensers,
    customer_supplied_imaging,
    labor_days,
    labor_misc_amt,
    distance_miles,
    include_mid_material,
    mid_brand,
    include_mid_labor,
    mid_labor_amt,
    tax_rate,
    gp_rate,
    company_key,
    cfg,
):
    """v13 imaging-only quote. Returns a dict with shape similar to compute_quote
    but tuned for the simpler imaging-only flow."""
    items_not_included = []
    imaging_amt = 0
    if not customer_supplied_imaging:
        prices = cfg["brand_imaging"].get(brand_name, {})
        if dispensers in prices:
            imaging_amt = prices[dispensers]
        else:
            items_not_included.append(
                f"Brand imaging price for {brand_name} at {dispensers} dispensers "
                f"is not on file in this configurator — verify before quoting."
            )
    shipping = cfg["shipping_handling"] if (not customer_supplied_imaging) else 0

    base_labor = labor_days * cfg["labor_daily_rate"]
    marked_up_labor = base_labor * (1 + gp_rate)
    threshold = cfg.get("distance_threshold_miles", 60)
    if distance_miles is not None and distance_miles > threshold:
        travel_adder_retail = labor_days * cfg.get("distance_adder_per_install_day", 0)
        travel_adder_cost = labor_days * cfg.get(
            "distance_adder_per_install_day_cost", 0
        )
    else:
        travel_adder_retail = 0
        travel_adder_cost = 0

    imaging_install = marked_up_labor + labor_misc_amt + travel_adder_retail

    mid_material = 0
    if include_mid_material:
        if mid_brand and mid_brand in cfg["mid_prices"]:
            mid_material = cfg["mid_prices"][mid_brand]
        else:
            items_not_included.append(
                f"Price sign material for {mid_brand} not in cheat sheet."
            )
    mid_labor_total = mid_labor_amt if include_mid_labor else 0

    material_total = imaging_amt + shipping + mid_material
    installation_total = imaging_install + mid_labor_total
    tax = material_total * tax_rate
    final = material_total + installation_total + tax

    # Profitability (best-effort)
    base_labor_cost = labor_days * cfg["labor_daily_rate_cost"]
    shipping_cost = (
        cfg["shipping_handling_cost"] if (not customer_supplied_imaging) else 0
    )
    brand_imaging_cost = imaging_amt  # pure pass-through
    mid_material_cost = (
        cfg["mid_prices_cost"].get(mid_brand, mid_material)
        if include_mid_material and mid_brand
        else 0
    )
    true_cost = (
        brand_imaging_cost
        + shipping_cost
        + base_labor_cost
        + labor_misc_amt
        + mid_material_cost
        + mid_labor_total
        + travel_adder_cost
    )
    revenue = material_total + installation_total
    profit = revenue - true_cost

    return {
        "imaging_amt": imaging_amt,
        "shipping": shipping,
        "imaging_install": imaging_install,
        "imaging_install_labor_marked": marked_up_labor,
        "imaging_install_misc": labor_misc_amt,
        "imaging_install_travel": travel_adder_retail,
        "mid_material": mid_material,
        "mid_labor": mid_labor_total,
        "labor_days": labor_days,
        "material_total": material_total,
        "installation_total": installation_total,
        "tax": tax,
        "final": final,
        "items_not_included": items_not_included,
        "revenue": revenue,
        "true_cost": true_cost,
        "profit": profit,
    }


# ════════════════════════════════════════════════════════════════════════════
# Proposal data builders (mode-aware)
# ════════════════════════════════════════════════════════════════════════════


def _canopy_data_block(q_canopy, r):
    """Build a single canopy dict for proposal_writer (used in both single and
    double modes). q_canopy = the per-canopy input dict; r = compute_quote result."""
    return {
        "fuel_type": q_canopy["fuel_type"],
        "type": q_canopy["canopy_type"],
        "branded": q_canopy["branded"],
        "brand_name": q_canopy["brand_name"] or "",
        "customer_supplied_imaging": q_canopy["customer_supplied_imaging"],
        "double_col": q_canopy["double_col"],
        "dispensers": q_canopy["dispensers"],
        "columns": r["columns"],
        "width": q_canopy["width"],
        "depth": q_canopy["depth"],
        "left_overhang": q_canopy["left_overhang"],
        "right_overhang": q_canopy["right_overhang"],
        "spacing": q_canopy["spacing"],
        "labor_days": r["labor_days"],
        "light_type": q_canopy["light_type"],
        "light_count": r["light_count"],
        "flood_light_count": r["flood_light_count"],
        "pricing": {
            "gp_rate": q_canopy["gp_rate"],
            "acm": r["acm"],
            "steel": r["steel"],
            "decking": r["decking"],
            "lights": r["lights"],
            "flood_lights": r["flood_lights"],
            "misc": r["misc"],
            "steel_secondary": r["steel_secondary"],
            "shipping": r["shipping"],
            "brand_imaging": r["brand_imaging"],
            "base_labor": r["base_labor"],
            "branded_labor_add": r["branded_labor_add"],
            "two_disp_labor_add": r["two_disp_labor_add"],
            "travel_adder_retail": r["travel_adder_retail"],
            "items_not_included": r["items_not_included"],
        },
    }


def _build_proposal_data(q):
    """Mode-aware proposal data builder."""
    mode = q.get("quote_mode", "single_canopy")
    base = {
        "company_key": q["company_key"],
        "quote_number": q["quote_number"],
        "quote_date": q["quote_date"],
        "customer": {
            "company": q["cust_company"],
            "name": q["cust_name"],
            "phone": q["cust_phone"],
            "email": q["cust_email"],
            "street": q["cust_street"],
            "city": q["cust_city"],
            "state": q["cust_state"],
            "zip": q["cust_zip"],
        },
        "sales_person": {
            "name": q["sales_person"],
            "phone": q["sales_info"]["phone"],
            "email": q["sales_info"]["email"],
        },
    }
    if mode == "imaging_only":
        r = q["result"]
        base.update(
            {
                "mode": "imaging_only",
                "imaging": {
                    "brand": q["brand_name"],
                    "dispensers": q["dispensers"],
                    "customer_supplied_imaging": q["customer_supplied_imaging"],
                    "imaging_amt": r["imaging_amt"],
                    "shipping": r["shipping"],
                    "labor_days": r["labor_days"],
                    "labor_daily_rate": q["labor_daily_rate"],
                    "labor_misc_amt": q["labor_misc_amt"],
                    "travel_adder_retail": r["imaging_install_travel"],
                },
                "mid": {
                    "include_material": q["include_mid_material"],
                    "include_labor": q["include_mid_labor"],
                    "brand": q["mid_brand"],
                    "material_amt": r["mid_material"],
                    "labor_amt": r["mid_labor"],
                },
                "pricing": {"gp_rate": q["gp_rate"], "tax_rate": q["tax_rate"]},
                "shared": {
                    "mid": {
                        "include_material": q["include_mid_material"],
                        "include_labor": q["include_mid_labor"],
                        "brand": q["mid_brand"],
                        "material_amt": r["mid_material"],
                        "labor_amt": r["mid_labor"],
                    },
                    "demo": {"include": False, "retail": 0, "cost": 0},
                    "tax_rate": q["tax_rate"],
                    "items_not_included": r.get("items_not_included", []),
                },
            }
        )
        return base
    if mode == "double_canopy":
        canopies = []
        items_not_inc = []
        for c, r in zip(q["canopies"], q["results"]):
            canopies.append(_canopy_data_block(c, r))
            items_not_inc += r["items_not_included"]
        base.update(
            {
                "mode": "double_canopy",
                "canopies": canopies,
                "shared": {
                    "mid": {
                        "include_material": q["include_mid_material"],
                        "include_labor": q["include_mid_labor"],
                        "brand": q["mid_brand"],
                        "material_amt": q["mid_material_amt"],
                        "labor_amt": q["mid_labor_amt"],
                    },
                    "demo": {
                        "include": q["include_demo"],
                        "retail": q["demo_retail"],
                        "cost": q["demo_cost"],
                    },
                    "tax_rate": q["tax_rate"],
                    "items_not_included": items_not_inc,
                },
            }
        )
        return base
    # single_canopy (v12-shape compatible)
    c = q["canopies"][0]
    r = q["results"][0]
    cd = _canopy_data_block(c, r)
    # Strip pricing into both top-level and into the canopy block (for v12 compat).
    base.update(
        {
            "mode": "single_canopy",
            "canopy": {k: v for k, v in cd.items() if k != "pricing"},
            "mid": {
                "include_material": q["include_mid_material"],
                "include_labor": q["include_mid_labor"],
                "brand": q["mid_brand"],
            },
            "demo": {
                "include": q["include_demo"],
                "retail": q["demo_retail"],
                "cost": q["demo_cost"],
            },
            "pricing": dict(
                cd["pricing"],
                tax_rate=q["tax_rate"],
                mid_material=r["mid_material"],
                mid_labor=r["mid_labor"],
            ),
        }
    )
    return base


def _docx_to_bytes(data):
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tf:
        tmp = tf.name
    try:
        build_proposal(data, tmp)
        with open(tmp, "rb") as f:
            return f.read()
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


def _next_quote_number():
    today = dt.date.today().strftime("%Y%m%d")
    prefix = f"Q-{today}-"
    max_seq = 0

    ws = _get_tracker_sheet()
    if ws is not None:
        try:
            quote_col = ws.col_values(2)[1:]
            for v in quote_col:
                if v and v.startswith(prefix):
                    try:
                        max_seq = max(max_seq, int(v.split("-")[-1]))
                    except (ValueError, IndexError):
                        pass
        except Exception:
            pass
    elif TRACKER_PATH.exists():
        try:
            with open(TRACKER_PATH, encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 2 and row[1].startswith(prefix):
                        try:
                            max_seq = max(max_seq, int(row[1].split("-")[-1]))
                        except (ValueError, IndexError):
                            pass
        except Exception:
            pass
    return f"{prefix}{max_seq + 1:03d}"


def _log_quote_to_tracker(q, *, grand_total, profit):
    row = [
        q["quote_date"],
        q["quote_number"],
        q["cust_company"],
        q["cust_city"],
        q["sales_person"],
        f"{grand_total:.2f}",
        f"{profit:.2f}",
    ]
    ws = _get_tracker_sheet()
    if ws is not None:
        ws.append_row(row, value_input_option="USER_ENTERED")
        return
    is_new = not TRACKER_PATH.exists()
    with open(TRACKER_PATH, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if is_new:
            w.writerow(TRACKER_COLUMNS)
        w.writerow(row)


def _street_name_for_filename(street):
    """v13 spec 3.7/7.4: take street name AFTER the street number, before the
    first comma. e.g. '5702 Brittmoore Road' -> 'Brittmoore Road'.
    '1234 N Main St Suite 200' -> 'N Main St Suite 200'."""
    if not street:
        return ""
    head = street.split(",")[0].strip()
    parts = head.split()
    if not parts:
        return ""
    # If first token is a number (or starts with one), drop it. Otherwise keep all.
    first = parts[0]
    if first[0].isdigit():
        return " ".join(parts[1:]).strip()
    return head


def _proposal_filename(q):
    base = f"Proposal_{q['quote_number']}"
    street = _street_name_for_filename(q.get("cust_street", ""))
    mode = q.get("quote_mode", "single_canopy")
    suffix_bits = []
    if street:
        suffix_bits.append(street)
    if mode == "imaging_only":
        suffix_bits.append("Imaging Only")
    if suffix_bits:
        base += f" ({' - '.join(suffix_bits)})"
    return base + ".docx"


# ════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Canopy Configurator", page_icon="⛽", layout="centered")

st.markdown(
    """
<style>
body { font-family: 'Segoe UI', sans-serif; }
.apec-banner { background: linear-gradient(135deg, #0d1b2a 0%, #1b3a5c 60%, #1a78c2 100%);
    border-radius: 12px; padding: 28px 24px 18px 24px; text-align: center; margin-bottom: 24px; }
.apec-logo-text { font-size: 2.6rem; font-weight: 900; letter-spacing: 3px; color: #f5a623;
    text-shadow: 2px 2px 6px rgba(0,0,0,0.5); }
.apec-sub { font-size: 1.05rem; color: #cde4f7; letter-spacing: 1px; margin-top: 4px; }
.section-title { font-size: 1.05rem; font-weight: 700; color: #1b3a5c;
    border-left: 4px solid #1a78c2; padding-left: 10px; margin: 22px 0 10px 0; }
.canopy-title { font-size: 1.15rem; font-weight: 800; color: #1b3a5c; background: #eef5fb;
    padding: 10px 16px; border-radius: 8px; margin: 24px 0 12px 0; border-left: 6px solid #1a78c2; }
.auto-readout { background: #f4f8fd; border: 1px dashed #b3d4f0; border-radius: 6px;
    padding: 8px 12px; font-size: 0.88rem; color: #1b3a5c; margin-top: 4px; }
.cost-card { background: #f4f8fd; border: 1px solid #b3d4f0; border-radius: 10px;
    padding: 16px 22px; margin: 12px 0; }
.cost-card .label { font-size: 0.85rem; color: #1b3a5c; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.5px; }
.cost-card .value { font-size: 1.4rem; color: #1b3a5c; font-weight: 700; }
.retail-card { background: linear-gradient(135deg, #1b3a5c 0%, #2a5a8c 100%);
    border-radius: 12px; padding: 24px 28px; margin: 14px 0; color: #fff; }
.retail-card .label { font-size: 1rem; color: #cde4f7; font-weight: 700;
    text-transform: uppercase; letter-spacing: 1px; }
.retail-card .value { font-size: 2.2rem; font-weight: 900; color: #f5a623; }
.summary-row { padding: 6px 0; border-bottom: 1px dotted #ccd9e8; font-size: 0.92rem; }
.summary-row .k { color: #6a7a8c; }
.summary-row .v { color: #1b3a5c; font-weight: 600; }
.disclaimer { background: #fff8e1; border-left: 4px solid #f5a623; padding: 12px 16px;
    font-size: 0.85rem; color: #5a4a1a; margin-top: 18px; border-radius: 6px; }
</style>
""",
    unsafe_allow_html=True,
)

try:
    cfg = load_config()
except FileNotFoundError:
    st.error(f"⚠️ `canopy_config.xlsx` not found at {CONFIG_PATH}.")
    st.stop()

brand_company = st.radio(
    "Is this quote for APEC Canopies or GEO Canopies?",
    ["APEC Canopies", "GEO Canopies"],
    horizontal=True,
)
company_key = "APEC" if "APEC" in brand_company else "GEO"
headline = (
    "APEC Imaging & Canopies — Price Configurator"
    if company_key == "APEC"
    else "GEO Canopies — Price Configurator"
)

active_logo = APEC_LOGO if company_key == "APEC" else GEO_LOGO
if active_logo.exists():
    st.image(str(active_logo), width=260)
st.markdown(
    f'<div class="apec-banner">'
    f'<div class="apec-logo-text">{"APEC" if company_key == "APEC" else "GEO"}</div>'
    f'<div class="apec-sub">{headline}</div></div>',
    unsafe_allow_html=True,
)

# ── Customer Information ──────────────────────────────────────────────
st.markdown(
    '<div class="section-title">Customer Information</div>', unsafe_allow_html=True
)
ca, cb = st.columns(2)
with ca:
    cust_company = st.text_input("Company Name")
    cust_name = st.text_input("Customer Name")
    cust_phone = st.text_input("Cell Phone")
    cust_email = st.text_input("Email")
with cb:
    cust_street = st.text_input("Street Address")
    cust_city = st.text_input("City")
    cust_state = st.selectbox("State", US_STATES, index=US_STATES.index("GA"))
    cust_zip = st.text_input("Zip")

threshold = int(cfg.get("distance_threshold_miles", 60))
distance_miles = st.number_input(
    f"Distance from {company_key} HQ (miles)",
    min_value=0,
    value=0,
    step=1,
    help=(
        f"If over {threshold} miles, a travel adder is silently added to "
        f"installation. Customer does not see this on the proposal."
    ),
)

st.markdown('<div class="section-title">Sales Person</div>', unsafe_allow_html=True)
sales_person = st.selectbox("Sales Person", list(cfg["salespeople"].keys()))
sales_info = cfg["salespeople"][sales_person]
st.markdown(
    f'<div class="auto-readout"><b>Cell:</b> {sales_info["phone"]} &nbsp;|&nbsp; '
    f"<b>Email:</b> {sales_info['email']}</div>",
    unsafe_allow_html=True,
)

tax_rate_pct = st.number_input(
    "Tax rate (%)", 0.0, 20.0, float(cfg["tax_default_rate"]) * 100, 0.25
)
tax_rate = tax_rate_pct / 100.0

# ── v13 Quote Type selector ──────────────────────────────────────────
st.markdown('<div class="section-title">Quote Type</div>', unsafe_allow_html=True)
quote_type = st.radio(
    "What is this quote for?",
    ["Canopies", "Imaging Only"],
    horizontal=True,
    index=0,
    help=(
        "Imaging Only = re-imaging an existing canopy "
        "(no new canopy build). Canopies = full canopy "
        "build + branding + price sign as applicable."
    ),
)

# Predefined GP based on company. v12 -3% rule for >=6 dispensers and v13 -3%
# rule for canopy 2 are applied where appropriate.
predefined_gp = cfg["gp_apec"] if company_key == "APEC" else cfg["gp_geo"]
LARGE_CANOPY_GP_REDUCTION = 0.03


def _password_gate(label="Override GP %", *, default_gp, key_prefix=""):
    """Common GP override widget. Returns (gp_rate, gp_overridden_bool)."""
    gp = default_gp
    overridden = False
    if st.checkbox(label, value=False, key=f"{key_prefix}_ov_gp"):
        pwd = st.text_input("Password", type="password", key=f"{key_prefix}_pwd")
        if pwd == PASSWORD:
            gp_pct = st.number_input(
                "Custom GP (%)",
                0.0,
                200.0,
                default_gp * 100,
                1.0,
                key=f"{key_prefix}_gp_val",
            )
            gp = gp_pct / 100.0
            overridden = True
        elif pwd:
            st.error("Incorrect password — using standard GP.")
    return gp, overridden


def render_canopy_inputs(
    *, key_prefix, fuel_locked=None, gp_default=None, gp_label="Override GP %"
):
    """Render the per-canopy input UI. Returns dict of inputs.

    key_prefix — used to namespace streamlit widget keys (must be unique).
    fuel_locked — if 'Gas' or 'Diesel', skip the fuel selector and use that.
    gp_default — default GP for this canopy (already accounts for any reductions).
    """
    out = {}
    # Canopy height (16.5 only for now)
    height_choice = st.radio(
        "Canopy height",
        ["16.5 ft (standard)", "20 ft"],
        horizontal=True,
        index=0,
        key=f"{key_prefix}_height",
    )
    if height_choice == "20 ft":
        st.error("⚠️ This configurator does not yet support 20 ft canopies.")
        st.stop()

    if fuel_locked is None:
        fuel_type = st.radio(
            "Fuel type",
            ["Gas", "Diesel"],
            horizontal=True,
            index=0,
            key=f"{key_prefix}_fuel",
            help="Gas defaults: LO 10, spacing 28, RO 10, depth 24. "
            "Diesel defaults: LO 2, spacing 24, RO 2, depth 20. "
            "All overridable below.",
        )
    else:
        fuel_type = fuel_locked
        st.caption(f"Fuel type: **{fuel_type}** (locked in double-canopy mode)")
    out["fuel_type"] = fuel_type

    canopy_type = st.radio(
        "Canopy type",
        ["Dive-in", "Stacked", "In-line"],
        horizontal=True,
        index=0,
        key=f"{key_prefix}_ctype",
    )
    dispensers = int(
        st.number_input(
            "Number of dispensers",
            min_value=2,
            max_value=10,
            value=int(cfg["default_dispensers"]),
            step=1,
            key=f"{key_prefix}_disp",
        )
    )
    if canopy_type == "Stacked" and dispensers % 2 != 0:
        st.error("⚠️ Stacked canopies require an EVEN number of dispensers.")
        st.stop()
    double_col = False
    if canopy_type == "Dive-in":
        col_arr = st.radio(
            "Column arrangement",
            ["Single column", "Double column"],
            horizontal=True,
            index=0,
            key=f"{key_prefix}_colarr",
        )
        double_col = col_arr == "Double column"
    columns = column_count(canopy_type, dispensers, double_col)
    st.markdown(
        f'<div class="auto-readout">Columns: <b>{columns}</b> (auto-calculated)</div>',
        unsafe_allow_html=True,
    )

    out["canopy_type"] = canopy_type
    out["dispensers"] = dispensers
    out["double_col"] = double_col

    if fuel_type == "Gas":
        default_lo = float(cfg["gas_default_left_overhang_ft"])
        default_ro = float(cfg["gas_default_right_overhang_ft"])
        default_spacing = float(cfg["gas_default_dispenser_spacing_ft"])
    else:
        default_lo = float(cfg["diesel_default_left_overhang_ft"])
        default_ro = float(cfg["diesel_default_right_overhang_ft"])
        default_spacing = float(cfg["diesel_default_dispenser_spacing_ft"])

    oc1, oc2, oc3 = st.columns(3)
    with oc1:
        ov_lo = st.checkbox(
            f"Override LO (default {default_lo:g}′)",
            value=False,
            key=f"{key_prefix}_ov_lo",
        )
        left_overhang = (
            st.number_input(
                "Left Overhang (ft)", 0.0, 50.0, default_lo, 1.0, key=f"{key_prefix}_lo"
            )
            if ov_lo
            else default_lo
        )
    with oc2:
        ov_sp = st.checkbox(
            f"Override spacing (default {default_spacing:g}′)",
            value=False,
            key=f"{key_prefix}_ov_sp",
        )
        spacing = (
            st.number_input(
                "Dispenser spacing (ft)",
                1.0,
                100.0,
                default_spacing,
                1.0,
                key=f"{key_prefix}_sp",
            )
            if ov_sp
            else default_spacing
        )
    with oc3:
        ov_ro = st.checkbox(
            f"Override RO (default {default_ro:g}′)",
            value=False,
            key=f"{key_prefix}_ov_ro",
        )
        right_overhang = (
            st.number_input(
                "Right Overhang (ft)",
                0.0,
                50.0,
                default_ro,
                1.0,
                key=f"{key_prefix}_ro",
            )
            if ov_ro
            else default_ro
        )
    out["left_overhang"] = left_overhang
    out["right_overhang"] = right_overhang
    out["spacing"] = spacing

    auto_w, auto_d = auto_dimensions(
        canopy_type, dispensers, left_overhang, right_overhang, spacing, fuel_type, cfg
    )
    ov_w = st.checkbox(
        f"Override width (auto: {auto_w:g}′)", value=False, key=f"{key_prefix}_ov_w"
    )
    width = (
        st.number_input(
            "Width (ft)", 1.0, 500.0, float(auto_w), 1.0, key=f"{key_prefix}_w"
        )
        if ov_w
        else float(auto_w)
    )
    ov_d = st.checkbox(
        f"Override depth (auto: {auto_d:g}′)", value=False, key=f"{key_prefix}_ov_d"
    )
    depth = (
        st.number_input(
            "Depth (ft)", 1.0, 500.0, float(auto_d), 1.0, key=f"{key_prefix}_d"
        )
        if ov_d
        else float(auto_d)
    )
    out["width"] = width
    out["depth"] = depth
    labor_days_preview = cfg["labor_days"].get(int(dispensers), 0)
    st.markdown(
        f'<div class="auto-readout">Labor days: <b>{labor_days_preview}</b> '
        f"&nbsp;|&nbsp; Square footage: <b>{int(width * depth):,}</b></div>",
        unsafe_allow_html=True,
    )

    # ── Canopy Lights ──
    light_type = st.radio(
        "Light type",
        ["Standard", "LSI", "Customer-provided"],
        horizontal=True,
        index=0,
        key=f"{key_prefix}_lt",
        help=(
            f"Standard ${int(cfg['light_unit_price_standard'])}/light; "
            f"LSI ${int(cfg['light_unit_price_lsi'])}/light; "
            "Customer-provided $0. Flat-priced (no GP markup)."
        ),
    )
    default_lights = default_light_count(canopy_type, dispensers, double_col)
    standard_light_count = columns * 4
    ov_lights = st.checkbox(
        f"Override number of canopy lights (default {default_lights})",
        value=False,
        key=f"{key_prefix}_ov_lights",
    )
    if ov_lights:
        light_count = int(
            st.number_input(
                "Number of canopy lights",
                min_value=0,
                max_value=400,
                value=default_lights,
                step=1,
                key=f"{key_prefix}_lights",
            )
        )
    else:
        light_count = default_lights
    needs_reconfirm = light_count != standard_light_count
    light_reconfirmed = True
    if needs_reconfirm:
        light_reconfirmed = st.checkbox(
            f"⚠️ Confirm canopy light count of {light_count} — standard is "
            f"4 per column ({standard_light_count}). Tick to confirm.",
            value=False,
            key=f"{key_prefix}_light_reconfirm",
        )
    out["light_type"] = light_type
    out["light_count"] = light_count
    out["needs_light_reconfirm"] = needs_reconfirm
    out["light_reconfirmed"] = light_reconfirmed

    # Flood lights
    flood_default_rate = int(cfg["flood_light_unit_price"])
    flood_light_count = int(
        st.number_input(
            "Number of flood lights",
            min_value=0,
            max_value=400,
            value=0,
            step=1,
            key=f"{key_prefix}_flood",
            help=(f"Flat ${flood_default_rate}/light. Default 0."),
        )
    )
    flood_light_unit_retail = flood_default_rate
    if flood_light_count > 0:
        if st.checkbox(
            f"Override flood light rate (default ${flood_default_rate}/light)",
            value=False,
            key=f"{key_prefix}_ov_flood_rate",
        ):
            flood_light_unit_retail = int(
                st.number_input(
                    "Flood light rate ($/light)",
                    min_value=flood_default_rate,
                    value=flood_default_rate,
                    step=25,
                    key=f"{key_prefix}_flood_rate",
                    help="Can only be raised above the default rate.",
                )
            )
        st.markdown(
            f'<div class="auto-readout">Flood lights: <b>{flood_light_count}</b> '
            f"@ ${flood_light_unit_retail}/light &nbsp;=&nbsp; "
            f"<b>${flood_light_count * flood_light_unit_retail:,}</b></div>",
            unsafe_allow_html=True,
        )
    out["flood_light_count"] = flood_light_count
    out["flood_light_unit_retail"] = flood_light_unit_retail

    # ── Branded Site (per-canopy in v13) ──
    branded_choice = st.radio(
        "Is this a branded site?",
        ["No (Unbranded)", "Yes (Branded)"],
        horizontal=True,
        index=0,
        key=f"{key_prefix}_branded",
    )
    branded = branded_choice == "Yes (Branded)"
    brand_name = None
    customer_supplied_imaging = False
    if branded:
        brand_options = list(cfg["brand_imaging"].keys())
        brand_name = st.selectbox(
            "Brand", brand_options, index=0, key=f"{key_prefix}_brand"
        )
        supply_choice = st.radio(
            "Imaging material supplied by:",
            [f"{company_key} provides", "Customer provides"],
            horizontal=True,
            index=0,
            key=f"{key_prefix}_supply",
            help=(
                "Pick who supplies imaging material. Auto-flips to Customer "
                "if no price on file for the chosen brand."
            ),
        )
        customer_supplied_imaging = supply_choice == "Customer provides"
        has_imaging_price = bool(cfg["brand_imaging"].get(brand_name))
        if not customer_supplied_imaging and not has_imaging_price:
            customer_supplied_imaging = True
            st.warning(
                f"No imaging price on file for **{brand_name}** — "
                f"defaulting to **Customer provides**."
            )
    out["branded"] = branded
    out["brand_name"] = brand_name
    out["customer_supplied_imaging"] = customer_supplied_imaging

    # ── GP override (per-canopy when in double mode) ──
    if gp_default is None:
        gp_default = predefined_gp
    # Apply >=6 disp -3% rule (v12) — applies to predefined GP only.
    large_canopy = dispensers >= 6
    eff_default = (
        (gp_default - LARGE_CANOPY_GP_REDUCTION) if large_canopy else gp_default
    )
    gp_rate, gp_overridden = _password_gate(
        gp_label, default_gp=eff_default, key_prefix=key_prefix
    )
    if large_canopy and not gp_overridden:
        st.caption(
            f"📉 {dispensers} dispensers (≥ 6) — GP reduced 3 points "
            f"to {eff_default * 100:.0f}%."
        )
    out["gp_rate"] = gp_rate
    out["gp_overridden"] = gp_overridden
    out["natural_gp_default"] = eff_default
    return out


# ════════════════════════════════════════════════════════════════════════════
# IMAGING-ONLY FLOW
# ════════════════════════════════════════════════════════════════════════════

if quote_type == "Imaging Only":
    st.markdown(
        '<div class="section-title">Imaging Configuration</div>', unsafe_allow_html=True
    )
    brand_options = list(cfg["brand_imaging"].keys())
    brand_name = st.selectbox("Brand", brand_options, index=0, key="img_brand")
    dispensers = int(
        st.number_input(
            "Number of dispensers (drives imaging price)",
            min_value=2,
            max_value=10,
            value=int(cfg["default_dispensers"]),
            step=1,
            key="img_disp",
        )
    )
    supply_choice = st.radio(
        "Imaging material supplied by:",
        [f"{company_key} provides", "Customer provides"],
        horizontal=True,
        index=0,
        key="img_supply",
        help="Customer-supplied → no imaging material or shipping charge.",
    )
    customer_supplied_imaging = supply_choice == "Customer provides"
    has_imaging_price = bool(cfg["brand_imaging"].get(brand_name))
    if not customer_supplied_imaging and not has_imaging_price:
        customer_supplied_imaging = True
        st.warning(
            f"No imaging price on file for **{brand_name}** — "
            f"defaulting to **Customer provides**."
        )

    st.markdown(
        '<div class="section-title">Installation Labor</div>', unsafe_allow_html=True
    )
    labor_days = int(
        st.number_input(
            "Labor days (uses canopy labor formula)",
            min_value=1,
            max_value=60,
            value=5,
            step=1,
            key="img_days",
        )
    )
    labor_misc_amt = int(
        st.number_input(
            "Additional labor / misc expenses ($)",
            min_value=0,
            value=0,
            step=100,
            key="img_misc",
            help="User-input dollar amount. Added to installation labor (non-taxable).",
        )
    )
    st.caption(
        f"Daily labor rate: ${int(cfg['labor_daily_rate']):,}/day "
        f"(× labor days × (1 + GP) + misc + travel = installation labor; "
        "non-taxable)."
    )

    # MID 4-way (with USER-INPUT labor amount)
    st.markdown(
        '<div class="section-title">MID / Price Sign</div>', unsafe_allow_html=True
    )
    include_mid_material = (
        st.radio(
            "Include price sign **material**?",
            ["No", "Yes"],
            horizontal=True,
            index=0,
            key="img_mid_mat",
        )
        == "Yes"
    )
    mid_brand = None
    if include_mid_material:
        mid_options = list(cfg["mid_prices"].keys())
        try:
            di = mid_options.index(brand_name)
        except ValueError:
            di = mid_options.index("Unbranded") if "Unbranded" in mid_options else 0
        mid_brand = st.selectbox(
            "Price sign brand", mid_options, index=di, key="img_mid_brand"
        )
    include_mid_labor = (
        st.radio(
            "Include price sign **installation labor**?",
            ["No", "Yes"],
            horizontal=True,
            index=0,
            key="img_mid_lab",
        )
        == "Yes"
    )
    mid_labor_amt = 0
    if include_mid_labor:
        # v13: imaging-only mode uses USER INPUT amount (no MID_Labor lookup).
        st.caption(
            "Imaging-only mode — enter the price sign installation amount manually."
        )
        mid_labor_amt = int(
            st.number_input(
                "Price sign installation amount ($)",
                min_value=0,
                value=8000,
                step=500,
                key="img_mid_labor_amt",
            )
        )
    if include_mid_labor and not include_mid_material:
        mid_options = list(cfg["mid_prices"].keys())
        try:
            di = mid_options.index(brand_name)
        except ValueError:
            di = mid_options.index("Unbranded") if "Unbranded" in mid_options else 0
        mid_brand = st.selectbox(
            "Price sign brand (for installation labeling)",
            mid_options,
            index=di,
            key="img_mid_brand_lab",
        )

    st.markdown(
        '<div class="section-title">GP Override (optional)</div>',
        unsafe_allow_html=True,
    )
    gp_rate, gp_overridden = _password_gate(
        "Override GP %", default_gp=predefined_gp, key_prefix="img"
    )

    st.markdown("<br>", unsafe_allow_html=True)
    calc = st.button(
        "⚡ Calculate Imaging Price",
        type="primary",
        use_container_width=True,
        key="img_calc_btn",
    )
    if calc:
        result = compute_imaging_only(
            brand_name=brand_name,
            dispensers=dispensers,
            customer_supplied_imaging=customer_supplied_imaging,
            labor_days=labor_days,
            labor_misc_amt=labor_misc_amt,
            distance_miles=distance_miles,
            include_mid_material=include_mid_material,
            mid_brand=mid_brand,
            include_mid_labor=include_mid_labor,
            mid_labor_amt=mid_labor_amt,
            tax_rate=tax_rate,
            gp_rate=gp_rate,
            company_key=company_key,
            cfg=cfg,
        )
        st.session_state["last_quote"] = {
            "quote_mode": "imaging_only",
            "result": result,
            "company_key": company_key,
            "brand_company": brand_company,
            "cust_company": cust_company,
            "cust_name": cust_name,
            "cust_phone": cust_phone,
            "cust_email": cust_email,
            "cust_street": cust_street,
            "cust_city": cust_city,
            "cust_state": cust_state,
            "cust_zip": cust_zip,
            "sales_person": sales_person,
            "sales_info": sales_info,
            "distance_miles": distance_miles,
            "brand_name": brand_name,
            "dispensers": dispensers,
            "customer_supplied_imaging": customer_supplied_imaging,
            "labor_days": labor_days,
            "labor_misc_amt": labor_misc_amt,
            "labor_daily_rate": cfg["labor_daily_rate"],
            "include_mid_material": include_mid_material,
            "include_mid_labor": include_mid_labor,
            "mid_brand": mid_brand,
            "mid_labor_amt": mid_labor_amt,
            "tax_rate": tax_rate,
            "tax_rate_pct": tax_rate_pct,
            "gp_rate": gp_rate,
            "quote_number": None,
            "quote_date": dt.date.today().strftime("%B %d, %Y"),
        }
        st.session_state.pop("proposal_bytes", None)
        st.session_state.pop("proposal_filename", None)
        st.session_state.pop("quote_logged", None)


# ════════════════════════════════════════════════════════════════════════════
# CANOPIES FLOW (single or double)
# ════════════════════════════════════════════════════════════════════════════

else:  # quote_type == "Canopies"
    st.markdown(
        '<div class="section-title">Canopy Configuration</div>', unsafe_allow_html=True
    )
    canopy_count_choice = st.radio(
        "Single or Double canopy quote?",
        ["Single Canopy", "Double Canopy (Gas + Diesel)"],
        horizontal=True,
        index=0,
        key="canopy_count_choice",
        help=(
            "Single = one canopy on this quote. Double = two canopies "
            "(typically a Truck Stop with a Gas canopy in front and a "
            "Diesel canopy in the back / on the side). Double mode "
            "configures both as fully independent canopies."
        ),
    )
    is_double = canopy_count_choice.startswith("Double")

    if not is_double:
        # ── SINGLE CANOPY MODE ──
        canopy_inputs = render_canopy_inputs(
            key_prefix="single",
            fuel_locked=None,
            gp_default=predefined_gp,
            gp_label="Override GP %",
        )
        canopy_inputs_list = [canopy_inputs]
    else:
        # ── DOUBLE CANOPY MODE ──
        st.markdown(
            '<div class="canopy-title">⛽ GAS CANOPY</div>', unsafe_allow_html=True
        )
        gas_inputs = render_canopy_inputs(
            key_prefix="gas",
            fuel_locked="Gas",
            gp_default=predefined_gp,
            gp_label="Override GP % (Gas canopy)",
        )
        st.markdown(
            '<div class="canopy-title">⛽ DIESEL CANOPY</div>', unsafe_allow_html=True
        )
        # Canopy 2 default GP = canopy 1 GP - 3% (per v13 spec 5.2e / 7.5b).
        canopy2_default = max(0.0, gas_inputs["gp_rate"] - DOUBLE_CANOPY_GP_REDUCTION)
        st.caption(
            f"💡 Default Diesel canopy GP = Gas canopy GP − 3% "
            f"= **{canopy2_default * 100:.0f}%**. Override below "
            f"with the GP password if a supervisor is making the quote."
        )
        diesel_inputs = render_canopy_inputs(
            key_prefix="diesel",
            fuel_locked="Diesel",
            gp_default=canopy2_default,
            gp_label="Override GP % (Diesel canopy)",
        )
        canopy_inputs_list = [gas_inputs, diesel_inputs]

    # ── Shared MID / Price Sign ──
    st.markdown(
        '<div class="section-title">MID / Price Sign (shared)</div>',
        unsafe_allow_html=True,
    )
    include_mid_material = (
        st.radio(
            "Include price sign **material**?",
            ["No", "Yes"],
            horizontal=True,
            index=0,
            key="shared_mid_mat",
        )
        == "Yes"
    )
    mid_brand = None
    mid_brand_options = list(cfg["mid_prices"].keys())
    if include_mid_material:
        # Suggest the first canopy's brand if it's branded
        suggested = (
            canopy_inputs_list[0].get("brand_name")
            if canopy_inputs_list[0].get("branded")
            else None
        )
        try:
            di = (
                mid_brand_options.index(suggested)
                if suggested
                else (
                    mid_brand_options.index("Unbranded")
                    if "Unbranded" in mid_brand_options
                    else 0
                )
            )
        except ValueError:
            di = (
                mid_brand_options.index("Unbranded")
                if "Unbranded" in mid_brand_options
                else 0
            )
        mid_brand = st.selectbox(
            "Price sign brand", mid_brand_options, index=di, key="shared_mid_brand"
        )
    include_mid_labor = (
        st.radio(
            "Include price sign **installation labor**?",
            ["No", "Yes"],
            horizontal=True,
            index=0,
            key="shared_mid_lab",
        )
        == "Yes"
    )
    mid_labor_amt = 0
    if include_mid_labor:
        default_mid_labor = int(
            cfg["mid_labor_by_company"].get(
                company_key, cfg["mid_labor_by_company"].get("APEC", 11000)
            )
        )
        st.caption(
            f"Default {company_key} price-sign installation labor: ${default_mid_labor:,}"
        )
        if st.checkbox(
            "Override price sign installation amount",
            value=False,
            key="shared_mid_lab_ov",
        ):
            mid_labor_amt = int(
                st.number_input(
                    "Custom price sign installation ($)",
                    min_value=0,
                    value=default_mid_labor,
                    step=500,
                    key="shared_mid_lab_amt",
                )
            )
        else:
            mid_labor_amt = default_mid_labor
    if include_mid_labor and not include_mid_material:
        suggested = (
            canopy_inputs_list[0].get("brand_name")
            if canopy_inputs_list[0].get("branded")
            else None
        )
        try:
            di = (
                mid_brand_options.index(suggested)
                if suggested
                else (
                    mid_brand_options.index("Unbranded")
                    if "Unbranded" in mid_brand_options
                    else 0
                )
            )
        except ValueError:
            di = (
                mid_brand_options.index("Unbranded")
                if "Unbranded" in mid_brand_options
                else 0
            )
        mid_brand = st.selectbox(
            "Price sign brand (for installation labeling)",
            mid_brand_options,
            index=di,
            key="shared_mid_brand_lab",
        )

    # ── Existing Canopy Removal (shared) ──
    st.markdown(
        '<div class="section-title">Existing Canopy Removal (shared)</div>',
        unsafe_allow_html=True,
    )
    include_demo = (
        st.radio(
            "Include existing canopy removal?",
            ["No", "Yes"],
            horizontal=True,
            index=0,
            key="shared_demo",
        )
        == "Yes"
    )
    demo_retail = 0
    demo_cost = 0
    if include_demo:
        default_demo_retail = int(cfg["demo_default_retail"])
        default_demo_cost = int(cfg["demo_default_cost"])
        st.caption(
            f"Defaults — Retail: ${default_demo_retail:,} / "
            f"Cost: ${default_demo_cost:,}"
        )
        if st.checkbox("Override removal price", value=False, key="shared_demo_ov"):
            rc1, rc2 = st.columns(2)
            with rc1:
                demo_retail = int(
                    st.number_input(
                        "Removal retail ($)",
                        min_value=0,
                        value=default_demo_retail,
                        step=500,
                        key="shared_demo_r",
                    )
                )
            with rc2:
                demo_cost = int(
                    st.number_input(
                        "Removal cost ($)",
                        min_value=0,
                        value=default_demo_cost,
                        step=500,
                        key="shared_demo_c",
                    )
                )
        else:
            demo_retail = default_demo_retail
            demo_cost = default_demo_cost

    st.markdown("<br>", unsafe_allow_html=True)
    calc = st.button(
        "⚡ Calculate Canopy Price",
        type="primary",
        use_container_width=True,
        key="canopies_calc_btn",
    )
    # Light-count reconfirm gate — applies to all canopies in the quote.
    if calc:
        for ci, cinp in enumerate(canopy_inputs_list):
            if cinp["needs_light_reconfirm"] and not cinp["light_reconfirmed"]:
                fuel_lbl = cinp["fuel_type"] if is_double else "this"
                st.error(
                    f"⚠️ Please tick the box to confirm the non-standard "
                    f"canopy light count for {fuel_lbl} canopy "
                    f"({cinp['light_count']} fixtures) before calculating."
                )
                calc = False
                break

    if calc:
        results = []
        for cinp in canopy_inputs_list:
            r = compute_quote(
                dispensers=cinp["dispensers"],
                canopy_type=cinp["canopy_type"],
                double_col=cinp["double_col"],
                fuel_type=cinp["fuel_type"],
                branded=cinp["branded"],
                brand_name=cinp["brand_name"],
                customer_supplied_imaging=cinp["customer_supplied_imaging"],
                include_mid_material=False,  # MID applied at quote level (shared)
                mid_brand=None,
                include_mid_labor=False,
                mid_labor_amt=0,
                left_overhang=cinp["left_overhang"],
                right_overhang=cinp["right_overhang"],
                spacing=cinp["spacing"],
                width=cinp["width"],
                depth=cinp["depth"],
                distance_miles=distance_miles,
                include_demo=False,  # demo applied at quote level (shared)
                demo_retail=0,
                demo_cost=0,
                tax_rate=tax_rate,
                gp_rate=cinp["gp_rate"],
                light_type=cinp["light_type"],
                light_count=cinp["light_count"],
                flood_light_count=cinp["flood_light_count"],
                flood_light_unit_retail=cinp["flood_light_unit_retail"],
                company_key=company_key,
                cfg=cfg,
            )
            results.append(r)

        st.session_state["last_quote"] = {
            "quote_mode": "double_canopy" if is_double else "single_canopy",
            "results": results,
            "canopies": canopy_inputs_list,
            "company_key": company_key,
            "brand_company": brand_company,
            "cust_company": cust_company,
            "cust_name": cust_name,
            "cust_phone": cust_phone,
            "cust_email": cust_email,
            "cust_street": cust_street,
            "cust_city": cust_city,
            "cust_state": cust_state,
            "cust_zip": cust_zip,
            "sales_person": sales_person,
            "sales_info": sales_info,
            "distance_miles": distance_miles,
            # Shared MID / demo for the proposal data builder
            "include_mid_material": include_mid_material,
            "include_mid_labor": include_mid_labor,
            "mid_brand": mid_brand,
            "mid_labor_amt": mid_labor_amt,
            # Compute MID material amount once at quote level
            "mid_material_amt": (
                cfg["mid_prices"].get(mid_brand, 0)
                if include_mid_material and mid_brand
                else 0
            ),
            "include_demo": include_demo,
            "demo_retail": demo_retail,
            "demo_cost": demo_cost,
            "tax_rate": tax_rate,
            "tax_rate_pct": tax_rate_pct,
            "quote_number": None,
            "quote_date": dt.date.today().strftime("%B %d, %Y"),
        }
        st.session_state.pop("proposal_bytes", None)
        st.session_state.pop("proposal_filename", None)
        st.session_state.pop("quote_logged", None)


# ════════════════════════════════════════════════════════════════════════════
# Display + Generate Proposal (mode-aware)
# ════════════════════════════════════════════════════════════════════════════

if "last_quote" in st.session_state:
    q = st.session_state["last_quote"]
    mode = q.get("quote_mode", "single_canopy")

    st.markdown('<div class="section-title">Quote Result</div>', unsafe_allow_html=True)

    # Compute display totals
    if mode == "imaging_only":
        r = q["result"]
        grand_total = r["final"]
        profit = r["profit"]
        retail_mat = r["material_total"]
        retail_lab = r["installation_total"]
        tax_amt = r["tax"]
        st.markdown(
            f'<div class="cost-card">'
            f'<div class="label">Imaging Cost (internal)</div>'
            f'<div class="value">Material: ${r["imaging_amt"] + r["shipping"] + r["mid_material"]:,.2f} '
            f"&nbsp;|&nbsp; Installation: ${r['imaging_install'] + r['mid_labor']:,.2f}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    else:
        # Sum across canopies; add shared MID + demo + tax.
        # Use proposal_writer's _combined_totals to stay in sync.
        from proposal_writer import _combined_totals

        prop_data = _build_proposal_data(q)
        totals = _combined_totals(prop_data)
        grand_total = totals["grand_total"]
        retail_mat = totals["material_total"]
        retail_lab = totals["installation_total"]
        tax_amt = totals["tax"]
        # Profit aggregate
        profit = sum(r["profit"] for r in q["results"])

    st.markdown(
        f'<div class="retail-card">'
        f'<div class="label">Retail Price</div>'
        f'<div class="value">${grand_total - tax_amt:,.2f}</div>'
        f'<div style="margin-top:8px;font-size:0.95rem;color:#cde4f7;">'
        f"Material: ${retail_mat:,.2f} &nbsp;|&nbsp; "
        f"Installation: ${retail_lab:,.2f} &nbsp;|&nbsp; "
        f"Tax ({q['tax_rate_pct']:.2f}%): ${tax_amt:,.2f}"
        f"</div>"
        f'<div style="margin-top:14px;font-size:1.1rem;font-weight:700;">'
        f"Final Price for Customer: ${grand_total:,.2f}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # Project summary (mode-specific)
    st.markdown(
        '<div class="section-title">Project Summary</div>', unsafe_allow_html=True
    )
    summary = [
        ("Quote for", q["brand_company"]),
        (
            "Quote mode",
            {
                "single_canopy": "Single canopy",
                "double_canopy": "Double canopy (Gas + Diesel)",
                "imaging_only": "Imaging only",
            }[mode],
        ),
        ("Customer", f"{q['cust_company']} — {q['cust_name']}"),
        ("Customer contact", f"{q['cust_phone']} | {q['cust_email']}"),
        (
            "Site address",
            f"{q['cust_street']}, {q['cust_city']}, {q['cust_state']} {q['cust_zip']}",
        ),
        (
            "Distance from HQ",
            f"{q['distance_miles']} miles"
            + (
                " (travel adder applied)"
                if q["distance_miles"] > cfg.get("distance_threshold_miles", 60)
                else ""
            ),
        ),
        (
            "Sales person",
            f"{q['sales_person']} ({q['sales_info']['phone']} | {q['sales_info']['email']})",
        ),
        ("Tax rate", f"{q['tax_rate_pct']:.2f}%"),
    ]
    if mode == "imaging_only":
        r = q["result"]
        summary += [
            ("Brand", q["brand_name"]),
            (
                "Imaging supplier",
                "Customer-supplied"
                if q["customer_supplied_imaging"]
                else f"{q['company_key']} provides",
            ),
            ("Dispensers", q["dispensers"]),
            ("Labor days", q["labor_days"]),
            ("Misc labor", f"${q['labor_misc_amt']:,}"),
            ("GP rate", f"{q['gp_rate'] * 100:.0f}%"),
        ]
    else:
        for ci, (cinp, r) in enumerate(zip(q["canopies"], q["results"])):
            tag = f"({cinp['fuel_type']})" if mode == "double_canopy" else ""
            summary.append(
                (
                    f"Canopy {ci + 1} {tag}".strip(),
                    f"{cinp['fuel_type']} {cinp['canopy_type']} "
                    f"/ {cinp['dispensers']} disp / "
                    f"{r['columns']} cols / GP {cinp['gp_rate'] * 100:.0f}%",
                )
            )
            summary.append(
                (
                    f"  └─ Branding",
                    (
                        f"Yes — {cinp['brand_name']}"
                        + (
                            " (Customer-supplied)"
                            if cinp["customer_supplied_imaging"]
                            else ""
                        )
                    )
                    if cinp["branded"]
                    else "No",
                )
            )
            summary.append(
                (
                    f"  └─ Lights",
                    f"{cinp['light_count']} {cinp['light_type']}"
                    + (
                        f" + {cinp['flood_light_count']} flood"
                        if cinp["flood_light_count"]
                        else ""
                    ),
                )
            )
    if mode != "imaging_only":
        if q.get("include_mid_material") or q.get("include_mid_labor"):
            ps = []
            if q.get("include_mid_material"):
                ps.append(f"material ({q.get('mid_brand') or '—'})")
            if q.get("include_mid_labor"):
                ps.append(f"installation (${q.get('mid_labor_amt', 0):,})")
            summary.append(("Price sign (shared)", " + ".join(ps)))
        else:
            summary.append(("Price sign (shared)", "Not included"))
        if q.get("include_demo"):
            summary.append(
                (
                    "Existing canopy removal (shared)",
                    f"Yes — ${q['demo_retail']:,} retail",
                )
            )
        else:
            summary.append(("Existing canopy removal (shared)", "No"))
    else:
        if q.get("include_mid_material") or q.get("include_mid_labor"):
            ps = []
            if q.get("include_mid_material"):
                ps.append(f"material ({q.get('mid_brand') or '—'})")
            if q.get("include_mid_labor"):
                ps.append(f"installation (${q.get('mid_labor_amt', 0):,})")
            summary.append(("Price sign", " + ".join(ps)))
        else:
            summary.append(("Price sign", "Not included"))

    for k, v in summary:
        st.markdown(
            f'<div class="summary-row"><span class="k">{k}:</span> '
            f'<span class="v">{v}</span></div>',
            unsafe_allow_html=True,
        )

    # Items not included (per-canopy or imaging)
    items_not_inc = []
    if mode == "imaging_only":
        items_not_inc = q["result"].get("items_not_included", [])
    else:
        for r in q["results"]:
            items_not_inc += r.get("items_not_included", [])
    if items_not_inc:
        st.markdown(
            '<div class="section-title">Items Not Included</div>',
            unsafe_allow_html=True,
        )
        for item in items_not_inc:
            st.markdown(f"- {item}")

    st.markdown(
        '<div class="disclaimer">'
        "<b>Disclaimer:</b> Price does not include permit, electrical work, "
        "canopy piers/footers, and bricking.<br><br>"
        "<i>This is an experimental configurator and not to be used in "
        "quoting real jobs at this time.</i>"
        "</div>",
        unsafe_allow_html=True,
    )

    # ── Generate Proposal ────────────────────────────────────────────────
    st.markdown(
        '<div class="section-title">Generate Proposal</div>', unsafe_allow_html=True
    )
    gen = st.button(
        "📄 Generate Word Proposal", use_container_width=True, key="gen_btn"
    )
    if gen:
        if not st.session_state.get("quote_logged"):
            q["quote_number"] = _next_quote_number()
            st.session_state["last_quote"] = q
            try:
                _log_quote_to_tracker(q, grand_total=grand_total, profit=profit)
                st.session_state["quote_logged"] = True
            except Exception as e:
                st.warning(
                    f"Quote tracker append failed (proposal still generated): {e}"
                )
        data = _build_proposal_data(q)
        try:
            st.session_state["proposal_bytes"] = _docx_to_bytes(data)
            st.session_state["proposal_filename"] = _proposal_filename(q)
            st.success(f"Proposal generated — {q['quote_number']}.")
        except Exception as e:
            st.error(f"Could not generate proposal: {e}")

    if "proposal_bytes" in st.session_state:
        st.download_button(
            "⬇️ Download Proposal (.docx)",
            data=st.session_state["proposal_bytes"],
            file_name=st.session_state["proposal_filename"],
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            type="primary",
            use_container_width=True,
        )


# ════════════════════════════════════════════════════════════════════════════
# Profitability Tracker (admin) — bottom of page, password-gated.
# ════════════════════════════════════════════════════════════════════════════

st.markdown("---")
with st.expander("🔒 Profitability Tracker (admin)"):
    admin_pw = st.text_input("Admin password", type="password", key="admin_pw")
    if admin_pw == ADMIN_PASSWORD:
        ws = _get_tracker_sheet()
        rows = []
        source_label = ""
        try:
            if ws is not None:
                rows = ws.get_all_values()
                source_label = "Google Sheets"
                st.markdown(
                    f"**Source:** [Open tracker in Google Sheets ↗]"
                    f"(https://docs.google.com/spreadsheets/d/{SHEET_ID})"
                )
            elif TRACKER_PATH.exists():
                with open(TRACKER_PATH, encoding="utf-8") as _f:
                    rows = list(csv.reader(_f))
                source_label = "Local CSV (fallback — Google Sheets not configured)"
                st.markdown(f"**Source:** {source_label}")
        except Exception as e:
            st.error(f"Tracker read error: {e}")
            rows = []

        if len(rows) >= 2:
            header, data_rows = rows[0], rows[1:]
            display_rows = list(reversed(data_rows))
            table_data = [dict(zip(header, r)) for r in display_rows]
            st.markdown(f"**{len(data_rows)} quote(s) logged.**")
            st.dataframe(table_data, use_container_width=True, hide_index=True)
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(header)
            for r_ in data_rows:
                w.writerow(r_)
            st.download_button(
                "⬇️ Download Tracker (.csv)",
                data=buf.getvalue().encode("utf-8"),
                file_name=f"quote_tracker_{dt.date.today()}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        else:
            st.info(
                "No quotes logged yet — tracker will populate after the "
                "first proposal is generated."
            )
    elif admin_pw:
        st.error("Incorrect admin password.")
