"""
profitability_report.py  —  Canopy Configurator internal profitability report.

Builds a line-by-line Excel (.xlsx) cost/retail/profit breakdown for a quote,
parallel to the Word proposal but INTERNAL (shows true costs).

Public API
----------
build_profitability_report(q, output_path) -> output_path
report_bytes(q) -> bytes            (for Streamlit download_button)
line_items(q)  -> list[dict]        (flat item rows; used for tie-out tests)

`q` is the same dict the app stores in st.session_state["last_quote"].

Design
------
* Cost & Retail are written as VALUES (editable inputs, blue) sourced from the
  pricing engine. $ Profit / Margin % / Markup % / subtotals / tax / grand
  totals are Excel FORMULAS, so editing any cost or retail recalculates live.
* A hidden category column (H) tags each item row "MAT" (taxable material) or
  "LAB" (installation). Tax is charged on material only; SUMIF keeps the
  material/labor split and grand totals robust regardless of which lines exist.
* Ties out to the proposal grand total: material_total + installation_total +
  tax  ==  customer 'final'.
"""

import io
import datetime as dt
from decimal import Decimal as _Dec, ROUND_HALF_UP as _RHU
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


def _round2(x):
    """Round to cents, half-away-from-zero (matches Excel ROUND)."""
    return float(_Dec(x).quantize(_Dec("0.01"), rounding=_RHU))

# ── palette ────────────────────────────────────────────────────────────────
NAVY      = "1F3864"
NAVY_MID  = "2E4D7B"
STEEL     = "8EAADB"
LIGHT     = "D9E1F2"
LIGHTER   = "EAEFF9"
SUBFILL   = "BDD7EE"
GREYHEAD  = "404040"
RED       = "C00000"
GREEN     = "375623"
GREENFILL = "C6EFCE"
REDFILL   = "FFC7CE"
INPUT_BLUE = "0000CC"
BLACK     = "000000"
WHITEX    = "FFFFFF"

FONT = "Arial"
MONEY = '$#,##0;($#,##0);-'
PCT   = '0.0%'
THIN  = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# column map (1-based)
C_ITEM, C_COST, C_RETAIL, C_PROFIT, C_MARGIN, C_MARKUP, C_GAP, C_CAT = 1, 2, 3, 4, 5, 6, 7, 8
NCOLS_VISIBLE = 6


# ════════════════════════════════════════════════════════════════════════════
# Data model: flatten a quote into item rows
# ════════════════════════════════════════════════════════════════════════════

def _canopy_label(mode, cinp, idx):
    if mode == "double_canopy":
        return f'{cinp.get("fuel_type", "").upper()} CANOPY'
    return "CANOPY"


def _canopy_items(r, cinp, canopy_label):
    """Item rows for one canopy result `r` + its input dict `cinp`."""
    gp = r.get("gp_rate", 0)
    out = []

    def add(section, label, cost, retail, cat, keep=False):
        if not keep and abs(cost) < 1e-6 and abs(retail) < 1e-6:
            return
        out.append(dict(canopy=canopy_label, section=section, label=label,
                        cost=round(cost, 2), retail=round(retail, 2), cat=cat))

    M = "MATERIAL"; S = "SUB-MATERIAL"; L = "LABOR / INSTALLATION"
    brand = (cinp.get("brand_name") or "").strip()
    # --- core (GP-marked) material ---
    add(M, "ACM fascia panels",            r["acm_cost"],     r["acm"]     * (1 + gp), "MAT", keep=True)
    add(M, f'Steel — primary frame ({cinp.get("dispensers","?")} sets)',
                                           r["steel_cost"],   r["steel"]   * (1 + gp), "MAT", keep=True)
    add(M, f'Decking ({r.get("columns","?")} columns)',
                                           r["decking_cost"], r["decking"] * (1 + gp), "MAT", keep=True)
    add(M, "MISC (hardware & consumables)", r["misc_cost"],   r["misc"]    * (1 + gp), "MAT", keep=True)
    # --- sub-material (flat / pass-through add-ons) ---
    add(S, f'Canopy lights — {r.get("light_type","")} ×{r.get("light_count",0)}',
                                           r["lights_cost"],        r["lights"],        "MAT")
    add(S, f'Flood lights ×{r.get("flood_light_count",0)}',
                                           r["flood_lights_cost"],  r["flood_lights"],  "MAT")
    add(S, "Steel — secondary (2nd set)",  r["steel_secondary_cost"], r["steel_secondary"], "MAT")
    add(S, f'Brand imaging — {brand or "brand"}',
                                           r["brand_imaging_cost"], r["brand_imaging"], "MAT")
    add(S, "Shipping & handling",          r["shipping_cost"],      r["shipping"],      "MAT")
    # --- labor / installation ---
    add(L, f'Installation labor ({r.get("labor_days","?")} days)',
                                           r["base_labor_cost"],        r["base_labor"] * (1 + gp), "LAB", keep=True)
    add(L, "Branded imaging labor adder",  r["branded_labor_add_cost"], r["branded_labor_add"],     "LAB")
    add(L, "2-dispenser labor adder",      r["two_disp_labor_add_cost"], r["two_disp_labor_add"],   "LAB")
    add(L, "Travel / distance adder",      r["travel_adder_cost"],      r["travel_adder_retail"],   "LAB")
    return out


def _shared_items(q):
    out = []
    SH = "SHARED ITEMS"
    if q.get("include_mid_material"):
        out.append(dict(canopy="SHARED", section=SH,
                        label=f'Price sign material — {(q.get("mid_brand") or "").strip()}'.rstrip(" —"),
                        cost=round(q.get("mid_material_cost_amt", q.get("mid_material_amt", 0)), 2),
                        retail=round(q.get("mid_material_amt", 0), 2), cat="MAT"))
    if q.get("include_mid_labor"):
        amt = round(q.get("mid_labor_amt", 0), 2)
        out.append(dict(canopy="SHARED", section=SH, label="Price sign installation",
                        cost=amt, retail=amt, cat="LAB"))
    if q.get("include_demo"):
        out.append(dict(canopy="SHARED", section=SH, label="Existing canopy removal",
                        cost=round(q.get("demo_cost", 0), 2),
                        retail=round(q.get("demo_retail", 0), 2), cat="LAB"))
    return out


def _imaging_items(q):
    r = q["result"]; gp = r.get("gp_rate", 0); out = []
    M = "MATERIAL"; L = "LABOR / INSTALLATION"; SH = "SHARED ITEMS"
    brand = (q.get("brand_name") or "").strip()
    cust = q.get("customer_supplied_imaging")
    suffix = " (customer-supplied)" if cust else ""

    def add(canopy, section, label, cost, retail, cat, keep=False):
        if not keep and abs(cost) < 1e-6 and abs(retail) < 1e-6:
            return
        out.append(dict(canopy=canopy, section=section, label=label,
                        cost=round(cost, 2), retail=round(retail, 2), cat=cat))

    add("IMAGING", M, f'Brand imaging — {brand or "brand"}{suffix}',
        r.get("imaging_amt_cost", r.get("imaging_amt", 0)), r.get("imaging_amt", 0), "MAT", keep=True)
    add("IMAGING", M, "Shipping & handling", r.get("shipping_cost", 0), r.get("shipping", 0), "MAT")
    add("IMAGING", L, f'Installation labor ({r.get("labor_days","?")} days)',
        r.get("base_labor_cost", 0), r.get("imaging_install_labor_marked", 0), "LAB", keep=True)
    add("IMAGING", L, "Additional labor / misc", r.get("labor_misc_cost", 0), r.get("imaging_install_misc", 0), "LAB")
    add("IMAGING", L, "Travel / distance adder", r.get("travel_adder_cost", 0), r.get("imaging_install_travel", 0), "LAB")
    # shared MID (lives inside the imaging result)
    if r.get("mid_material", 0):
        add("SHARED", SH, f'Price sign material — {(q.get("mid_brand") or "").strip()}'.rstrip(" —"),
            r.get("mid_material_cost", 0), r.get("mid_material", 0), "MAT", keep=True)
    if r.get("mid_labor", 0):
        add("SHARED", SH, "Price sign installation",
            r.get("mid_labor_cost", 0), r.get("mid_labor", 0), "LAB", keep=True)
    return out


def line_items(q):
    """Flat list of item rows (no subtotals) — also used for tie-out tests."""
    mode = q.get("quote_mode", "single_canopy")
    if mode == "imaging_only":
        return _imaging_items(q)
    items = []
    for i, (r, cinp) in enumerate(zip(q["results"], q["canopies"])):
        items += _canopy_items(r, cinp, _canopy_label(mode, cinp, i))
    items += _shared_items(q)
    return items


def computed_totals(q):
    """Pure-python totals (for verification). Mirrors the engine."""
    items = line_items(q)
    mat_retail = sum(it["retail"] for it in items if it["cat"] == "MAT")
    lab_retail = sum(it["retail"] for it in items if it["cat"] == "LAB")
    total_cost = sum(it["cost"] for it in items)
    tax_rate = q.get("tax_rate", 0.0)
    tax = _round2(mat_retail * tax_rate)
    grand = round(mat_retail + lab_retail + tax, 2)
    return dict(material_retail=round(mat_retail, 2), labor_retail=round(lab_retail, 2),
                pretax_retail=round(mat_retail + lab_retail, 2), total_cost=round(total_cost, 2),
                tax=tax, grand_total=grand, profit=round(mat_retail + lab_retail - total_cost, 2))


# ════════════════════════════════════════════════════════════════════════════
# Rendering
# ════════════════════════════════════════════════════════════════════════════

def _set(ws, row, col, value, *, bold=False, size=10, color=BLACK, fill=None,
         align=None, fmt=None, italic=False, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align or ("left" if col == C_ITEM else "right"),
                            vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = BORDER_ALL
    return c


def _profit_formulas(ws, row):
    """Add $Profit / Margin% / Markup% formulas for a data/subtotal/total row."""
    cost = f'{get_column_letter(C_COST)}{row}'
    ret  = f'{get_column_letter(C_RETAIL)}{row}'
    _set(ws, row, C_PROFIT, f'={ret}-{cost}', fmt=MONEY)
    _set(ws, row, C_MARGIN, f'=IFERROR(({ret}-{cost})/{ret},"")', fmt=PCT)
    _set(ws, row, C_MARKUP, f'=IFERROR(({ret}-{cost})/{cost},"")', fmt=PCT)


def _band(ws, row, text, *, fill=NAVY_MID, color=WHITEX, size=10):
    ws.merge_cells(start_row=row, start_column=C_ITEM, end_row=row, end_column=C_MARKUP)
    _set(ws, row, C_ITEM, text, bold=True, size=size, color=color, fill=fill, align="left")
    for col in range(C_COST, C_MARKUP + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row=row, column=col).border = BORDER_ALL
    ws.row_dimensions[row].height = 18


def _total_label(ws, row, text, *, fill, color=WHITEX, size=11):
    """Label-only banded row (does NOT merge value columns, so B–F stay writable)."""
    _set(ws, row, C_ITEM, text, bold=True, size=size, color=color, fill=fill, align="left")
    ws.row_dimensions[row].height = 20


def build_profitability_report(q, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Profitability"
    ws.sheet_view.showGridLines = False

    widths = {C_ITEM: 46, C_COST: 15, C_RETAIL: 15, C_PROFIT: 15,
              C_MARGIN: 12, C_MARKUP: 12, C_GAP: 2, C_CAT: 6}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.column_dimensions[get_column_letter(C_CAT)].hidden = True

    mode = q.get("quote_mode", "single_canopy")
    mode_label = {"single_canopy": "Single canopy",
                  "double_canopy": "Double canopy (Gas + Diesel)",
                  "imaging_only": "Imaging only"}.get(mode, mode)

    row = 1
    # ── Title ──
    ws.merge_cells(start_row=row, start_column=C_ITEM, end_row=row, end_column=C_MARKUP)
    _set(ws, row, C_ITEM, "CANOPY PROFITABILITY REPORT", bold=True, size=16,
         color=WHITEX, fill=NAVY, align="left", border=False)
    for col in range(C_COST, C_MARKUP + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[row].height = 26
    row += 1
    ws.merge_cells(start_row=row, start_column=C_ITEM, end_row=row, end_column=C_MARKUP)
    _set(ws, row, C_ITEM, "INTERNAL — CONTAINS TRUE COSTS — NOT FOR CUSTOMER DISTRIBUTION",
         bold=True, size=9, color=RED, align="left", border=False)
    row += 2

    # ── Meta block ──
    cust_addr = ", ".join(x for x in [q.get("cust_street", ""),
                f'{q.get("cust_city","")}, {q.get("cust_state","")} {q.get("cust_zip","")}'.strip(" ,")]
                if x.strip(" ,"))
    meta = [
        ("Quote number", q.get("quote_number") or "—", "Date", q.get("quote_date") or dt.date.today().strftime("%B %d, %Y")),
        ("Customer", q.get("cust_company") or q.get("cust_name") or "—", "Contact", q.get("cust_name") or "—"),
        ("Site address", cust_addr or "—", "Sales person", q.get("sales_person") or "—"),
        ("Quote type", mode_label, "Prepared by", q.get("brand_company") or q.get("company_key") or "—"),
    ]
    taxrate_row = None
    for L1, V1, L2, V2 in meta:
        _set(ws, row, C_ITEM, L1, bold=True, align="left", border=False)
        ws.merge_cells(start_row=row, start_column=C_COST, end_row=row, end_column=C_PROFIT)
        _set(ws, row, C_COST, V1, align="left", border=False)
        _set(ws, row, C_MARGIN, L2, bold=True, align="left", border=False)
        _set(ws, row, C_MARKUP, V2, align="left", border=False)
        row += 1
    # tax rate (editable assumption)
    _set(ws, row, C_ITEM, "Sales tax rate", bold=True, align="left", border=False)
    taxrate_row = row
    _set(ws, row, C_COST, q.get("tax_rate", 0.0), color=INPUT_BLUE, align="left",
         fmt='0.00%', border=False)
    _set(ws, row, C_MARGIN, "(editable)", italic=True, size=8, color=GREYHEAD, align="left", border=False)
    row += 2

    # ── Table header ──
    headers = ["Line item", "Cost", "Retail", "$ Profit", "Margin %", "Markup %"]
    for col, h in enumerate(headers, start=1):
        _set(ws, row, col, h, bold=True, size=10, color=WHITEX, fill=GREYHEAD,
             align="left" if col == C_ITEM else "center")
    ws.row_dimensions[row].height = 18
    header_row = row
    row += 1

    items = line_items(q)

    # order canopies as they appear, then SHARED last
    canopy_order = []
    for it in items:
        if it["canopy"] not in canopy_order:
            canopy_order.append(it["canopy"])
    # ensure SHARED is last
    if "SHARED" in canopy_order:
        canopy_order = [c for c in canopy_order if c != "SHARED"] + ["SHARED"]

    section_order = ["MATERIAL", "SUB-MATERIAL", "LABOR / INSTALLATION", "SHARED ITEMS"]
    data_first = row
    subtotal_rows = []          # (row, kind) for material/labor totals
    first_item_row = None
    last_item_row = None

    for canopy in canopy_order:
        c_items = [it for it in items if it["canopy"] == canopy]
        if not c_items:
            continue
        # canopy band (skip for single 'CANOPY' if you like; keep for clarity)
        band_text = canopy if canopy != "SHARED" else "SHARED ITEMS (across canopies)"
        _band(ws, row, band_text, fill=NAVY_MID)
        row += 1
        present_sections = [s for s in section_order if any(it["section"] == s for it in c_items)]
        for section in present_sections:
            s_items = [it for it in c_items if it["section"] == section]
            if not s_items:
                continue
            # section sub-band
            _set(ws, row, C_ITEM, section, bold=True, size=9, color=NAVY,
                 fill=LIGHT, align="left")
            for col in range(C_COST, C_MARKUP + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=LIGHT)
                ws.cell(row=row, column=col).border = BORDER_ALL
            row += 1
            sec_start = row
            for it in s_items:
                _set(ws, row, C_ITEM, it["label"], align="left")
                _set(ws, row, C_COST, it["cost"], color=INPUT_BLUE, fmt=MONEY)
                _set(ws, row, C_RETAIL, it["retail"], color=INPUT_BLUE, fmt=MONEY)
                _profit_formulas(ws, row)
                ws.cell(row=row, column=C_CAT, value=it["cat"])
                if first_item_row is None:
                    first_item_row = row
                last_item_row = row
                row += 1
            sec_end = row - 1
            # section subtotal
            _set(ws, row, C_ITEM, f"   {section} subtotal", bold=True, align="left", fill=LIGHTER)
            _set(ws, row, C_COST, f'=SUM({get_column_letter(C_COST)}{sec_start}:{get_column_letter(C_COST)}{sec_end})',
                 bold=True, fmt=MONEY, fill=LIGHTER)
            _set(ws, row, C_RETAIL, f'=SUM({get_column_letter(C_RETAIL)}{sec_start}:{get_column_letter(C_RETAIL)}{sec_end})',
                 bold=True, fmt=MONEY, fill=LIGHTER)
            _profit_formulas(ws, row)
            for col in (C_PROFIT, C_MARGIN, C_MARKUP):
                ws.cell(row=row, column=col).font = Font(name=FONT, bold=True, size=10)
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=LIGHTER)
            row += 1
        row += 0  # tight

    last_data_row = row - 1

    # ── Totals block ──
    row += 1
    cl_cost = get_column_letter(C_COST); cl_ret = get_column_letter(C_RETAIL)
    cl_cat = get_column_letter(C_CAT)
    rng = lambda c: f'{get_column_letter(c)}{first_item_row}:{get_column_letter(c)}{last_item_row}'
    catrng = f'{cl_cat}{first_item_row}:{cl_cat}{last_item_row}'

    # Pre-tax grand total (SUMIF over item rows only)
    _total_label(ws, row, "GRAND TOTAL — before tax", fill=NAVY)
    pretax_row = row
    _set(ws, row, C_COST, f'=SUMIF({catrng},"MAT",{rng(C_COST)})+SUMIF({catrng},"LAB",{rng(C_COST)})',
         bold=True, color=WHITEX, fill=NAVY, fmt=MONEY)
    _set(ws, row, C_RETAIL, f'=SUMIF({catrng},"MAT",{rng(C_RETAIL)})+SUMIF({catrng},"LAB",{rng(C_RETAIL)})',
         bold=True, color=WHITEX, fill=NAVY, fmt=MONEY)
    _set(ws, row, C_PROFIT, f'={cl_ret}{row}-{cl_cost}{row}', bold=True, color=WHITEX, fill=NAVY, fmt=MONEY)
    _set(ws, row, C_MARGIN, f'=IFERROR(({cl_ret}{row}-{cl_cost}{row})/{cl_ret}{row},"")', bold=True, color=WHITEX, fill=NAVY, fmt=PCT)
    _set(ws, row, C_MARKUP, f'=IFERROR(({cl_ret}{row}-{cl_cost}{row})/{cl_cost}{row},"")', bold=True, color=WHITEX, fill=NAVY, fmt=PCT)
    row += 1

    # Material subtotal & Labor subtotal (informational, drives tax)
    _set(ws, row, C_ITEM, "   of which — taxable material", align="left", italic=True)
    matsum_row = row
    _set(ws, row, C_RETAIL, f'=SUMIF({catrng},"MAT",{rng(C_RETAIL)})', fmt=MONEY, italic=True)
    _set(ws, row, C_COST, f'=SUMIF({catrng},"MAT",{rng(C_COST)})', fmt=MONEY, italic=True)
    for col in (C_PROFIT, C_MARGIN, C_MARKUP):
        ws.cell(row=row, column=col).border = BORDER_ALL
    row += 1
    _set(ws, row, C_ITEM, "   of which — installation / labor", align="left", italic=True)
    _set(ws, row, C_RETAIL, f'=SUMIF({catrng},"LAB",{rng(C_RETAIL)})', fmt=MONEY, italic=True)
    _set(ws, row, C_COST, f'=SUMIF({catrng},"LAB",{rng(C_COST)})', fmt=MONEY, italic=True)
    for col in (C_PROFIT, C_MARGIN, C_MARKUP):
        ws.cell(row=row, column=col).border = BORDER_ALL
    row += 1

    # Tax row (material * rate)
    _set(ws, row, C_ITEM, "Sales tax (pass-through, no profit)", align="left")
    tax_row = row
    _set(ws, row, C_RETAIL, f'=ROUND({cl_ret}{matsum_row}*{cl_cost}{taxrate_row},2)', fmt=MONEY)
    _set(ws, row, C_COST, f'={cl_ret}{tax_row}', fmt=MONEY)   # cost==retail -> 0 profit
    _profit_formulas(ws, row)
    row += 1

    # Final grand total (with tax)
    _total_label(ws, row, "GRAND TOTAL — customer price (incl. tax)", fill=GREEN)
    final_row = row
    _set(ws, row, C_COST, f'={cl_cost}{pretax_row}+{cl_cost}{tax_row}', bold=True, color=WHITEX, fill=GREEN, fmt=MONEY)
    _set(ws, row, C_RETAIL, f'={cl_ret}{pretax_row}+{cl_ret}{tax_row}', bold=True, color=WHITEX, fill=GREEN, fmt=MONEY)
    _set(ws, row, C_PROFIT, f'={cl_ret}{row}-{cl_cost}{row}', bold=True, color=WHITEX, fill=GREEN, fmt=MONEY)
    _set(ws, row, C_MARGIN, f'=IFERROR(({cl_ret}{row}-{cl_cost}{row})/{cl_ret}{row},"")', bold=True, color=WHITEX, fill=GREEN, fmt=PCT)
    _set(ws, row, C_MARKUP, f'=IFERROR(({cl_ret}{row}-{cl_cost}{row})/{cl_cost}{row},"")', bold=True, color=WHITEX, fill=GREEN, fmt=PCT)
    row += 2

    # Footnote
    ws.merge_cells(start_row=row, start_column=C_ITEM, end_row=row, end_column=C_MARKUP)
    _set(ws, row, C_ITEM,
         "Margin % = Profit ÷ Retail (share of sale price kept).   "
         "Markup % = Profit ÷ Cost (uplift over what we pay).   "
         "Blue cells are editable — change a Cost or Retail to test procurement / pricing scenarios.",
         italic=True, size=8, color=GREYHEAD, align="left", border=False, wrap=True)
    ws.row_dimensions[row].height = 28

    # ── Conditional formatting: colour-scale the $ Profit + Margin% on item rows ──
    if first_item_row and last_item_row >= first_item_row:
        prof_range = f'{get_column_letter(C_PROFIT)}{first_item_row}:{get_column_letter(C_PROFIT)}{last_item_row}'
        marg_range = f'{get_column_letter(C_MARGIN)}{first_item_row}:{get_column_letter(C_MARGIN)}{last_item_row}'
        ws.conditional_formatting.add(prof_range, ColorScaleRule(
            start_type='min', start_color=REDFILL,
            mid_type='percentile', mid_value=50, mid_color=WHITEX,
            end_type='max', end_color=GREENFILL))
        ws.conditional_formatting.add(marg_range, ColorScaleRule(
            start_type='num', start_value=0, start_color=REDFILL,
            mid_type='num', mid_value=0.25, mid_color="FFEB9C",
            end_type='num', end_value=0.5, end_color=GREENFILL))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    return output_path


def report_bytes(q):
    buf = io.BytesIO()
    build_profitability_report(q, buf)
    return buf.getvalue()
